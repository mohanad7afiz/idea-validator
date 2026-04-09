"""Generate a sample workbook for idea-validator demos.

This produces `sample-ideas.xlsx` with 3 generic business ideas in the
17-column idea-vault format. Anyone cloning idea-validator can use this
sample to try the `/validate-idea` skill without having their own workbook.
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill


EXPECTED_COLUMNS = [
    "Idea Name",
    "Description",
    "Target Audience",
    "Revenue Model",
    "Pricing Range",
    "Monthly Income Potential",
    "Setup Cost",
    "Time to Build",
    "Maintenance Level",
    "Pros",
    "Cons",
    "Competitors",
    "Market Demand",
    "Difficulty",
    "Proven Examples",
    "Business Setup Required",
    "Why This Problem Hurts",
]


SAMPLE_IDEAS = [
    {
        "Idea Name": "Slack Reminder Bot",
        "Description": "A Slack bot that lets users set natural-language reminders in channels or DMs, with recurring schedules, snooze, and team delegation.",
        "Target Audience": "Distributed engineering and product teams using Slack as their primary comms tool",
        "Revenue Model": "Freemium SaaS — free for up to 10 reminders/month, paid tiers for unlimited and team features",
        "Pricing Range": "$0 free; $5/user/month Pro; $15/user/month Team",
        "Monthly Income Potential": "$500-$5,000 early; $20,000+ with 500 paying teams",
        "Setup Cost": "$50 (Slack app registration, Heroku hobby, domain)",
        "Time to Build": "1-2 weeks",
        "Maintenance Level": "Low",
        "Pros": "Ships fast; clear pain point; sticky once installed; organic team-level spread",
        "Cons": "Slack-only; Slack itself has reminders (commodity); pricing sensitivity",
        "Competitors": "Slack native /remind, Howdy, Geekbot",
        "Market Demand": "Medium",
        "Difficulty": "Low",
        "Proven Examples": "Geekbot ($1M+ ARR), Standuply",
        "Business Setup Required": "None — MoR platform (Lemon Squeezy/Paddle)",
        "Why This Problem Hurts": "Slack's native /remind is bare-bones; teams need recurring, delegatable, context-aware reminders and currently glue together 3 tools",
    },
    {
        "Idea Name": "AI Cat Photo Enhancer",
        "Description": "Upload a cat photo, get back an enhanced version: brighter, sharper, with optional stylization (watercolor, oil painting, studio portrait).",
        "Target Audience": "Cat owners sharing photos on Instagram, TikTok, and cat-dedicated communities",
        "Revenue Model": "Pay-per-pack — $5 for 20 enhanced photos, $15 subscription for unlimited",
        "Pricing Range": "$5 one-off; $15/mo subscription",
        "Monthly Income Potential": "$2,000-$20,000+ with social virality",
        "Setup Cost": "$300 (GPU via Replicate, domain, Stripe)",
        "Time to Build": "2-3 weeks",
        "Maintenance Level": "Medium",
        "Pros": "Emotional connection drives sharing; low technical bar; clear output quality delta",
        "Cons": "Cat photo niche (vs. general pets); seasonal traffic; thin moat",
        "Competitors": "Remini, PhotoRoom, general photo enhancers",
        "Market Demand": "High (niche but passionate)",
        "Difficulty": "Low",
        "Proven Examples": "Photo enhancers generally do well; cat-specific verticals less proven",
        "Business Setup Required": "None — MoR platform",
        "Why This Problem Hurts": "Cat owners take thousands of photos; most look mediocre; generic photo apps don't understand cat faces well",
    },
    {
        "Idea Name": "API Status Dashboard",
        "Description": "A self-hosted dashboard that pings your API endpoints on a schedule and shows uptime, latency, and error rates in a clean web UI.",
        "Target Audience": "Solo developers and small engineering teams who need basic API health visibility without the complexity of Datadog",
        "Revenue Model": "Open source with paid hosted version; $10/month for 10 endpoints, $30/month for 50",
        "Pricing Range": "$0 self-hosted; $10-30/month hosted",
        "Monthly Income Potential": "$1,000-$10,000 with modest adoption",
        "Setup Cost": "$100 (domain, hosting)",
        "Time to Build": "2-4 weeks",
        "Maintenance Level": "Medium",
        "Pros": "Clear ROI; small teams hate overbuilt monitoring; open source earns trust",
        "Cons": "Crowded monitoring space; commoditized by free tools; sustaining both OSS and paid is hard",
        "Competitors": "UptimeRobot, BetterUptime, Datadog (overkill), Pingdom",
        "Market Demand": "Medium",
        "Difficulty": "Medium",
        "Proven Examples": "BetterUptime ($1M+ ARR), UptimeRobot (acquired)",
        "Business Setup Required": "None — MoR platform",
        "Why This Problem Hurts": "Teams find out about API outages from customer complaints; existing tools are either too bare or too enterprise",
    },
]


def build_workbook(output_path: Path) -> None:
    """Generate sample-ideas.xlsx at the given path."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Sample Ideas"

    # Header row
    for col_idx, col_name in enumerate(EXPECTED_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")

    # Data rows
    for row_idx, idea in enumerate(SAMPLE_IDEAS, start=2):
        for col_idx, col_name in enumerate(EXPECTED_COLUMNS, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=idea[col_name])

    # Column widths — rough, just so it's readable
    for col_idx in range(1, len(EXPECTED_COLUMNS) + 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25

    wb.save(output_path)


def main() -> None:
    out = Path(__file__).parent / "sample-ideas.xlsx"
    build_workbook(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
