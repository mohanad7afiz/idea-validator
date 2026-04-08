"""Tests for the sample workbook generator."""
from pathlib import Path
import tempfile

import openpyxl
import pytest

from generate_sample import build_workbook, SAMPLE_IDEAS, EXPECTED_COLUMNS


def test_expected_columns_has_17_fields():
    assert len(EXPECTED_COLUMNS) == 17
    assert EXPECTED_COLUMNS[0] == "Idea Name"


def test_sample_ideas_has_three_entries():
    assert len(SAMPLE_IDEAS) == 3


def test_sample_ideas_each_have_all_columns():
    for idea in SAMPLE_IDEAS:
        for col in EXPECTED_COLUMNS:
            assert col in idea, f"missing column {col!r} in idea {idea.get('Idea Name')!r}"


def test_build_workbook_writes_xlsx(tmp_path: Path):
    out = tmp_path / "sample-ideas.xlsx"
    build_workbook(out)
    assert out.exists()


def test_workbook_has_one_sheet_with_correct_headers(tmp_path: Path):
    out = tmp_path / "sample-ideas.xlsx"
    build_workbook(out)
    wb = openpyxl.load_workbook(out)
    assert len(wb.sheetnames) == 1
    sheet = wb[wb.sheetnames[0]]
    header_row = [cell.value for cell in sheet[1]]
    assert header_row == EXPECTED_COLUMNS


def test_workbook_has_three_data_rows(tmp_path: Path):
    out = tmp_path / "sample-ideas.xlsx"
    build_workbook(out)
    wb = openpyxl.load_workbook(out)
    sheet = wb[wb.sheetnames[0]]
    data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 3
    # Each row must have 17 non-None values
    for i, row in enumerate(data_rows):
        assert len(row) == 17, f"row {i} has {len(row)} cells"
        assert all(cell is not None for cell in row), f"row {i} has None cells"
